"""
parser.py â€” Excel payroll report parser
Handles both stock report format (Payroll Analysis Updated.xlsx)
Parses all 7 sheets, normalizes dates, carries down facility names,
excludes Grand Total rows, handles stacked metric blocks in PPDs.
"""

import re
import json
import openpyxl
from datetime import datetime, date
from dateutil.parser import parse as dateutil_parse


# â”€â”€ Date Normalization â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

EXCEL_EPOCH = datetime(1899, 12, 30)

KNOWN_SERIALS = {
    46104: "2026-03-23",
    46109: "2026-03-28",
    46112: "2026-03-31",
    46116: "2026-04-04",
    46123: "2026-04-11",
    46130: "2026-04-18",
}


def normalize_date(val) -> str | None:
    """Convert any date-like value to ISO YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        serial = int(val)
        if serial in KNOWN_SERIALS:
            return KNOWN_SERIALS[serial]
        # Generic Excel serial conversion
        try:
            dt = EXCEL_EPOCH + __import__('datetime').timedelta(days=serial)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        # Already ISO
        if re.match(r'^\d{4}-\d{2}-\d{2}', val):
            return val[:10]
        # M/D/YYYY
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', val)
        if m:
            return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        try:
            return dateutil_parse(val).strftime("%Y-%m-%d")
        except Exception:
            return None
    return None


def date_label(iso: str) -> str:
    """Convert ISO date to display label like 3/28/2026."""
    try:
        dt = datetime.strptime(iso, "%Y-%m-%d")
        return f"{dt.month}/{dt.day}/{dt.year}"
    except Exception:
        return iso


# ── Helper: treat empty string as None ────────────────────────────────────────

def is_empty(val):
    return val is None or str(val).strip() == ""

def carry_down(prev, current):
    """Return current if not blank, else prev."""
    if is_empty(current) or str(current).strip().lower() == "none":
        return prev
    return str(current).strip()


# ── GRAND TOTAL / TOTAL row detection ─────────────────────────────────────────

TOTAL_PATTERNS = re.compile(
    r'^(grand\s*total|total)$', re.IGNORECASE
)


def is_total_row(val) -> bool:
    if is_empty(val):
        return False
    return bool(TOTAL_PATTERNS.match(str(val).strip()))


def to_float(val) -> float | None:
    if is_empty(val):
        return None
    s = str(val).strip()
    if s == '-':
        return 0.0
    # Remove currency, spaces, commas, and % signs
    s = re.sub(r'[$,%\s]', '', s)
    # Handle accounting format parentheses (123) -> -123
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    try:
        f = float(s)
        return f
    except (ValueError, TypeError):
        return None


# ── Sheet parsers ─────────────────────────────────────────────────────────────


def parse_analysis_by_acq_group(ws) -> dict:
    """
    Sheet: Analysis by Acq Group
    Structure: Multiple metric blocks stacked vertically.
    Each block: metric label row, blank row, header row (Period End Date | groups…), data rows.
    """
    rows = list(ws.iter_rows(values_only=True))
    result = {}  # metric_name -> { acq_group -> { period_date -> value } }

    current_metric = None
    header_row = None

    for row in rows:
        first_cell = str(row[0]).strip() if not is_empty(row[0]) else ""
        # Detect metric label rows (non-date, non-empty, not "Period End Date")
        if first_cell and first_cell.lower() != "period end date" and not normalize_date(row[0]):
            current_metric = first_cell
            header_row = None
            continue

        # Detect header row "Period End Date | group1 | group2 …"
        if first_cell.lower() == "period end date":
            header_row = [str(c).strip() if not is_empty(c) else None for c in row]
            continue

        # Data row: first cell is a date
        dt = normalize_date(row[0])
        if dt and current_metric and header_row:
            if current_metric not in result:
                result[current_metric] = {}
            for col_idx, col_name in enumerate(header_row[1:], start=1):
                if col_name and col_name.lower() not in ("none", ""):
                    val = to_float(row[col_idx]) if col_idx < len(row) else None
                    if col_name not in result[current_metric]:
                        result[current_metric][col_name] = {}
                    result[current_metric][col_name][dt] = val

    return result


def parse_analysis_by_region(ws) -> dict:
    """
    Sheet: Analysis by Region
    Same stacked-block structure but with region columns.
    Region headers contain facility counts like "Central - 13 Facilities".
    """
    rows = list(ws.iter_rows(values_only=True))
    result = {}  # metric_name -> { region -> { period_date -> value } }
    region_info = {}  # region_name -> facility_count

    current_metric = None
    header_row = None

    for row in rows:
        first_cell = str(row[0]).strip() if not is_empty(row[0]) else ""

        if first_cell and first_cell.lower() != "period end date" and not normalize_date(row[0]):
            current_metric = first_cell
            header_row = None
            continue

        if first_cell.lower() == "period end date":
            header_row = []
            for c in row:
                if not is_empty(c):
                    s = str(c).strip()
                    # Extract region name and facility count
                    m = re.match(r'^(.+?)\s*-\s*(\d+)\s*Facilit', s, re.IGNORECASE)
                    if m:
                        region_name = m.group(1).strip()
                        count = int(m.group(2))
                        region_info[region_name] = count
                        header_row.append(region_name)
                    else:
                        header_row.append(s)
                else:
                    header_row.append(None)
            continue

        dt = normalize_date(row[0])
        if dt and current_metric and header_row:
            if current_metric not in result:
                result[current_metric] = {}
            for col_idx, col_name in enumerate(header_row[1:], start=1):
                if col_name and col_name.lower() not in ("none", "", "total"):
                    val = to_float(row[col_idx]) if col_idx < len(row) else None
                    if col_name not in result[current_metric]:
                        result[current_metric][col_name] = {}
                    result[current_metric][col_name][dt] = val

    return {"metrics": result, "region_info": region_info}


def parse_ot_by_pay_period(ws) -> list[dict]:
    """
    Sheet: OT by Pay Period
    Row 3: group header (OT Dollars, OT Hours, OT % of Gross, OT % of Worked Hours)
    Row 4: column header with pay period dates as strings
    Rows 5+: detail rows with carried-down subgroup/facility
    CRITICAL: Stop carry-down on Grand Total rows.
    Returns list of dicts.
    """
    rows = list(ws.iter_rows(values_only=True))
    col_header_idx = None
    data_start = None

    for i, row in enumerate(rows):
        # Allow empty strings for missing cells
        if is_empty(row[0]) and not is_empty(row[5]) and str(row[5]).strip() in ('Employee Name', 'Employee'):
            col_header_idx = i
            data_start = i + 1
            break
        if not is_empty(row[0]) and str(row[0]).strip() in ('Subgroup Name', 'Subgroup'):
            col_header_idx = i
            data_start = i + 1
            break

    if col_header_idx is None or data_start is None:
        col_header_idx = 3 if len(rows) > 3 else 0
        data_start = col_header_idx + 1

    group_labels_row = rows[col_header_idx - 1] if col_header_idx > 0 else []
    date_header_row = rows[col_header_idx] if col_header_idx < len(rows) else []

    # Build column map: col_idx -> (group, period_date)
    col_map = {}
    current_group = None
    for idx, val in enumerate(group_labels_row):
        if not is_empty(val):
            current_group = str(val).strip()
        if idx >= 6:  # data starts at col 6 (0-indexed)
            dt = normalize_date(date_header_row[idx]) if idx < len(date_header_row) else None
            if dt:
                col_map[idx] = (current_group, dt)

    records = []
    last_subgroup = None
    last_facility = None
    last_department = None

    for row in rows[data_start:]:
        if all(is_empty(v) for v in row):
            continue

        subgroup_raw = row[0] if len(row) > 0 else None
        facility_raw = row[1] if len(row) > 1 else None
        dept_raw = row[2] if len(row) > 2 else None
        position_raw = row[4] if len(row) > 4 else None
        employee_raw = row[5] if len(row) > 5 else None

        # Detect Grand Total / Total rows — do NOT carry them
        if is_total_row(facility_raw) or is_total_row(subgroup_raw):
            last_subgroup = None
            last_facility = None
            last_department = None
            continue

        # Carry-down logic
        if not is_empty(subgroup_raw):
            last_subgroup = str(subgroup_raw).strip()
        if not is_empty(facility_raw):
            last_facility = str(facility_raw).strip()
        if not is_empty(dept_raw):
            last_department = str(dept_raw).strip()

        if is_empty(last_facility) or is_empty(employee_raw):
            continue

        employee = str(employee_raw).strip()
        position = str(position_raw).strip() if not is_empty(position_raw) else None

        # Extract metric values per period
        for col_idx, (group_label, period_date) in col_map.items():
            val = to_float(row[col_idx]) if col_idx < len(row) else None
            if val is None:
                continue
            records.append({
                "acquisition_group": last_subgroup,
                "facility_name": last_facility,
                "department": last_department,
                "position": position,
                "employee_name": employee,
                "period_date": period_date,
                "metric_group": group_label,
                "value": val,
            })

    return records


def parse_top_ot_earners(ws) -> list[dict]:
    """
    Sheet: Top OT Earners
    Row 4: headers — Employee, Facility, Position, [OT Dollars dates…], [OT Hours dates…]
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row
    header_idx = None
    for i, row in enumerate(rows):
        if not is_empty(row[0]) and str(row[0]).strip() in ('Employee', 'Employee Name'):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 3

    header = rows[header_idx]
    # Build column map
    ot_dollar_cols = {}   # period_date -> col_idx
    ot_hour_cols = {}
    total_dollar_col = None
    total_hour_col = None

    group_row = rows[header_idx - 1] if header_idx > 0 else [None] * 20
    current_group = None
    for idx, val in enumerate(group_row):
        if not is_empty(val):
            current_group = str(val).strip()

    in_dollars = True
    for idx, val in enumerate(header[4:], start=4):
        dt = normalize_date(val)
        if dt:
            if in_dollars:
                ot_dollar_cols[dt] = idx
            else:
                ot_hour_cols[dt] = idx
        elif not is_empty(val):
            s = str(val).strip()
            if 'Total OT Dollar' in s or 'Total OT $' in s:
                total_dollar_col = idx
                in_dollars = False
            elif 'Total OT Hour' in s:
                total_hour_col = idx
                in_dollars = False
            elif 'OT Hours' in s:
                in_dollars = False

    records = []
    for row in rows[header_idx + 1:]:
        if all(is_empty(v) for v in row):
            continue
        employee = str(row[0]).strip() if not is_empty(row[0]) else None
        facility = str(row[1]).strip() if not is_empty(row[1]) else None
        position = str(row[2]).strip() if not is_empty(row[2]) else None
        if not employee:
            continue

        total_dollars = to_float(row[total_dollar_col]) if total_dollar_col and total_dollar_col < len(row) else None
        total_hours = to_float(row[total_hour_col]) if total_hour_col and total_hour_col < len(row) else None

        for dt, col in ot_dollar_cols.items():
            ot_d = to_float(row[col]) if col < len(row) else None
            ot_h_col = ot_hour_cols.get(dt)
            ot_h = to_float(row[ot_h_col]) if ot_h_col and ot_h_col < len(row) else None
            records.append({
                "employee_name": employee,
                "facility_name": facility,
                "position": position,
                "period_date": dt,
                "ot_dollars": ot_d,
                "ot_hours": ot_h,
                "total_ot_dollars": total_dollars,
                "total_ot_hours": total_hours,
            })

    return records


def parse_bonus_by_ppe(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if not is_empty(row[0]) and str(row[0]).strip() in ('Subgroup Name', 'Subgroup'):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 2

    header = rows[header_idx]
    date_cols = {}
    for idx, val in enumerate(header[6:], start=6):
        dt = normalize_date(val)
        if dt:
            date_cols[idx] = dt

    records = []
    last_subgroup = None
    last_facility = None
    last_bonus_type = None

    for row in rows[header_idx + 1:]:
        if all(is_empty(v) for v in row):
            continue
        if is_total_row(row[1]) or is_total_row(row[0]):
            continue

        sg = row[0]
        fac = row[1]
        bt = row[2]

        if not is_empty(sg):
            last_subgroup = str(sg).strip()
        if not is_empty(fac):
            last_facility = str(fac).strip()
        if not is_empty(bt):
            last_bonus_type = str(bt).strip()

        position = str(row[4]).strip() if not is_empty(row[4]) else None
        employee = str(row[5]).strip() if not is_empty(row[5]) else None

        if not last_facility or not employee:
            continue

        for col_idx, period_date in date_cols.items():
            val = to_float(row[col_idx]) if col_idx < len(row) else None
            if val is None:
                continue
            records.append({
                "acquisition_group": last_subgroup,
                "facility_name": last_facility,
                "bonus_type": last_bonus_type,
                "position": position,
                "employee_name": employee,
                "period_date": period_date,
                "bonus_dollars": val,
            })

    return records


def parse_bonus_by_type(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if not is_empty(row[0]) and str(row[0]).strip() in ('Bonus Type',):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 2

    header = rows[header_idx]
    date_cols = {}
    for idx, val in enumerate(header[6:], start=6):
        dt = normalize_date(val)
        if dt:
            date_cols[idx] = dt

    records = []
    last_bonus_type = None
    last_subgroup = None
    last_facility = None

    for row in rows[header_idx + 1:]:
        if all(is_empty(v) for v in row):
            continue
        if is_total_row(row[2]) or is_total_row(row[1]):
            continue

        bt = row[0]
        sg = row[1]
        fac = row[2]

        if not is_empty(bt):
            last_bonus_type = str(bt).strip()
        if not is_empty(sg):
            last_subgroup = str(sg).strip()
        if not is_empty(fac):
            last_facility = str(fac).strip()

        position = str(row[4]).strip() if len(row) > 4 and not is_empty(row[4]) else None
        employee = str(row[5]).strip() if len(row) > 5 and not is_empty(row[5]) else None

        if not last_facility or not employee:
            continue

        for col_idx, period_date in date_cols.items():
            val = to_float(row[col_idx]) if col_idx < len(row) else None
            if val is None:
                continue
            records.append({
                "acquisition_group": last_subgroup,
                "facility_name": last_facility,
                "bonus_type": last_bonus_type,
                "position": position,
                "employee_name": employee,
                "period_date": period_date,
                "bonus_dollars": val,
            })

    return records


def parse_ppds(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    records = []

    metric_blocks = {
        "Direct Care HPPD": None,
        "Direct Care PPD": None,
        "Overall Labor PPD": None
    }

    current_metric = None
    header_idx = None
    date_cols = {}

    for i, row in enumerate(rows):
        first_cell = str(row[0]).strip() if not is_empty(row[0]) else ""

        if first_cell in metric_blocks:
            current_metric = first_cell
            header_idx = None
            date_cols = {}
            continue

        if current_metric and (first_cell == "Subgroup" or (not is_empty(row[1]) and str(row[1]).strip() in ("Subgroup", "Subgroup Name"))):
            header_idx = i
            for idx, val in enumerate(row[4:], start=4):
                dt = normalize_date(val)
                if dt:
                    date_cols[idx] = dt
            continue

        if current_metric and header_idx is not None and i > header_idx:
            if all(is_empty(v) for v in row):
                current_metric = None
                continue

            if is_total_row(row[1]) or is_total_row(row[0]):
                continue

            sg = row[0]
            fac = row[1]

            subgroup = str(sg).strip() if not is_empty(sg) else None
            facility = str(fac).strip() if not is_empty(fac) else None

            if not facility:
                continue

            for col_idx, period_date in date_cols.items():
                val = to_float(row[col_idx]) if col_idx < len(row) else None
                if val is None:
                    continue
                records.append({
                    "metric_type": current_metric,
                    "acquisition_group": subgroup,
                    "facility_name": facility,
                    "period_date": period_date,
                    "value": val,
                })

    return records


# â”€â”€ Main parse entry point â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

REQUIRED_SHEETS = {
    "Analysis by Acq Group",
    "Analysis by Region",
    "OT by Pay Period",
    "Top OT Earners",
    "Bonus by PPE",
    "Bonus by Type",
    "PPDs",
}

class MockWorksheet:
    def __init__(self, data):
        self.data = data
        data_max_col = max((len(r) for r in data), default=0) if data else 0
        self.max_col = max(data_max_col, 50)

    def iter_rows(self, values_only=True):
        for row in self.data:
            row_len = len(row)
            if row_len < self.max_col:
                yield tuple(row) + (None,) * (self.max_col - row_len)
            else:
                yield tuple(row)

def parse_workbook_json(sheets_data: dict) -> dict:
    """
    Parse the stock payroll report directly from JSON lists (from browser SheetJS).
    Bypasses openpyxl entirely for speed and avoiding serverless timeouts.
    """
    sheet_names = set(sheets_data.keys())
    missing_sheets = REQUIRED_SHEETS - sheet_names
    
    issues = []
    if missing_sheets:
        for s in missing_sheets:
            issues.append({"severity": "HIGH", "category": "Missing Sheet",
                           "location": s, "description": f"Required sheet '{s}' not found in workbook."})

    result = {
        "sheets_found": list(sheet_names),
        "sheets_missing": list(missing_sheets),
        "issues": issues,
        "pay_periods": set(),
        "facilities": set(),
        "acquisition_groups": set(),
        "regions": set(),
    }

    if "Analysis by Acq Group" in sheet_names:
        ag_data = parse_analysis_by_acq_group(MockWorksheet(sheets_data["Analysis by Acq Group"]))
        result["acq_group_metrics"] = ag_data
        for metric_name, groups in ag_data.items():
            for group, periods in groups.items():
                if group.lower() != "total":
                    result["acquisition_groups"].add(group)
                for dt in periods:
                    result["pay_periods"].add(dt)
    else:
        result["acq_group_metrics"] = {}

    if "Analysis by Region" in sheet_names:
        region_data = parse_analysis_by_region(MockWorksheet(sheets_data["Analysis by Region"]))
        result["region_metrics"] = region_data["metrics"]
        result["region_info"] = region_data["region_info"]
        for metric_name, regions in region_data["metrics"].items():
            for region, periods in regions.items():
                result["regions"].add(region)
                for dt in periods:
                    result["pay_periods"].add(dt)
    else:
        result["region_metrics"] = {}
        result["region_info"] = {}

    if "OT by Pay Period" in sheet_names:
        ot_records = parse_ot_by_pay_period(MockWorksheet(sheets_data["OT by Pay Period"]))
        result["ot_detail"] = ot_records
        for r in ot_records:
            result["facilities"].add(r["facility_name"])
            result["pay_periods"].add(r["period_date"])
            if r["acquisition_group"]:
                result["acquisition_groups"].add(r["acquisition_group"])
    else:
        result["ot_detail"] = []

    if "Top OT Earners" in sheet_names:
        result["top_ot_earners"] = parse_top_ot_earners(MockWorksheet(sheets_data["Top OT Earners"]))
    else:
        result["top_ot_earners"] = []

    if "Bonus by PPE" in sheet_names:
        bonus_ppe = parse_bonus_by_ppe(MockWorksheet(sheets_data["Bonus by PPE"]))
        result["bonus_by_ppe"] = bonus_ppe
        for r in bonus_ppe:
            result["facilities"].add(r["facility_name"])
            result["pay_periods"].add(r["period_date"])
    else:
        result["bonus_by_ppe"] = []

    if "Bonus by Type" in sheet_names:
        result["bonus_by_type"] = parse_bonus_by_type(MockWorksheet(sheets_data["Bonus by Type"]))
    else:
        result["bonus_by_type"] = []

    if "PPDs" in sheet_names:
        ppd_records = parse_ppds(MockWorksheet(sheets_data["PPDs"]))
        result["ppd_metrics"] = ppd_records
        for r in ppd_records:
            result["facilities"].add(r["facility_name"])
            result["pay_periods"].add(r["period_date"])
    else:
        result["ppd_metrics"] = []

    result["pay_periods"] = sorted(result["pay_periods"])
    result["facilities"] = sorted(result["facilities"])
    result["acquisition_groups"] = sorted(result["acquisition_groups"])
    result["regions"] = sorted(result["regions"])

    return result

def parse_workbook(filepath: str) -> dict:
    """
    Parse the stock payroll report.
    Returns structured dict with all parsed data + metadata.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = set(wb.sheetnames)
    missing_sheets = REQUIRED_SHEETS - sheet_names
    present_sheets = REQUIRED_SHEETS & sheet_names

    issues = []
    if missing_sheets:
        for s in missing_sheets:
            issues.append({"severity": "HIGH", "category": "Missing Sheet",
                           "location": s, "description": f"Required sheet '{s}' not found in workbook."})

    result = {
        "sheets_found": list(sheet_names),
        "sheets_missing": list(missing_sheets),
        "issues": issues,
        "pay_periods": set(),
        "facilities": set(),
        "acquisition_groups": set(),
        "regions": set(),
    }

    # â”€â”€ Analysis by Acq Group â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "Analysis by Acq Group" in sheet_names:
        ag_data = parse_analysis_by_acq_group(wb["Analysis by Acq Group"])
        result["acq_group_metrics"] = ag_data
        for metric_name, groups in ag_data.items():
            for group, periods in groups.items():
                if group.lower() != "total":
                    result["acquisition_groups"].add(group)
                for dt in periods:
                    result["pay_periods"].add(dt)
    else:
        result["acq_group_metrics"] = {}

    # â”€â”€ Analysis by Region â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "Analysis by Region" in sheet_names:
        region_data = parse_analysis_by_region(wb["Analysis by Region"])
        result["region_metrics"] = region_data["metrics"]
        result["region_info"] = region_data["region_info"]
        for metric_name, regions in region_data["metrics"].items():
            for region, periods in regions.items():
                result["regions"].add(region)
                for dt in periods:
                    result["pay_periods"].add(dt)
    else:
        result["region_metrics"] = {}
        result["region_info"] = {}

    # â”€â”€ OT by Pay Period â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "OT by Pay Period" in sheet_names:
        ot_records = parse_ot_by_pay_period(wb["OT by Pay Period"])
        result["ot_detail"] = ot_records
        for r in ot_records:
            result["facilities"].add(r["facility_name"])
            result["pay_periods"].add(r["period_date"])
            if r["acquisition_group"]:
                result["acquisition_groups"].add(r["acquisition_group"])
    else:
        result["ot_detail"] = []

    # â”€â”€ Top OT Earners â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "Top OT Earners" in sheet_names:
        result["top_ot_earners"] = parse_top_ot_earners(wb["Top OT Earners"])
    else:
        result["top_ot_earners"] = []

    # â”€â”€ Bonus by PPE â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "Bonus by PPE" in sheet_names:
        bonus_ppe = parse_bonus_by_ppe(wb["Bonus by PPE"])
        result["bonus_by_ppe"] = bonus_ppe
        for r in bonus_ppe:
            result["facilities"].add(r["facility_name"])
            result["pay_periods"].add(r["period_date"])
    else:
        result["bonus_by_ppe"] = []

    # â”€â”€ Bonus by Type â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "Bonus by Type" in sheet_names:
        result["bonus_by_type"] = parse_bonus_by_type(wb["Bonus by Type"])
    else:
        result["bonus_by_type"] = []

    # â”€â”€ PPDs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "PPDs" in sheet_names:
        ppd_records = parse_ppds(wb["PPDs"])
        result["ppd_metrics"] = ppd_records
        for r in ppd_records:
            result["facilities"].add(r["facility_name"])
            result["pay_periods"].add(r["period_date"])
    else:
        result["ppd_metrics"] = []

    # Convert sets to sorted lists
    result["pay_periods"] = sorted(result["pay_periods"])
    result["facilities"] = sorted(result["facilities"])
    result["acquisition_groups"] = sorted(result["acquisition_groups"])
    result["regions"] = sorted(result["regions"])

    return result


# â”€â”€ Payroll Schedule Group Detection â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def detect_payroll_schedule_groups(parsed: dict) -> dict:
    """
    Analyze which pay periods each facility appears in from OT/PPD data.
    Assign each facility to a payroll schedule group based on its pay period pattern.
    Returns: {facility_name -> {schedule_group, pay_periods}}
    """
    # Build facility -> set of periods from OT detail
    facility_periods = {}
    for r in parsed.get("ot_detail", []):
        fac = r["facility_name"]
        dt = r["period_date"]
        if fac not in facility_periods:
            facility_periods[fac] = set()
        facility_periods[fac].add(dt)

    # Also include PPD data
    for r in parsed.get("ppd_metrics", []):
        fac = r["facility_name"]
        dt = r["period_date"]
        if fac not in facility_periods:
            facility_periods[fac] = set()
        facility_periods[fac].add(dt)

    # Group facilities by their period pattern
    pattern_to_group = {}
    facility_groups = {}

    for fac, periods in facility_periods.items():
        pattern = tuple(sorted(periods))
        if pattern not in pattern_to_group:
            letter = chr(ord('A') + len(pattern_to_group))
            period_labels = ", ".join([date_label(p) for p in pattern])
            group_name = f"Group {letter} ({period_labels})"
            pattern_to_group[pattern] = group_name
        facility_groups[fac] = {
            "schedule_group": pattern_to_group[pattern],
            "pay_periods": sorted(periods),
        }

    return facility_groups


